"""
Yordamchi funksiyalar
"""

import hashlib
import hmac
import uuid
import qrcode
import os
from datetime import datetime, timedelta, time
from typing import List, Dict, Optional, Tuple
from io import BytesIO
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from config import Config
import pytz
import re

def generate_ticket_id(booking_id: int, court_id: int, date: datetime) -> str:
    """
    Bilet ID yaratish
    Format: TNS-YYYYMMDD-CRT{N}-XXXX
    """
    date_str = date.strftime("%Y%m%d")
    random_suffix = str(uuid.uuid4())[:4].upper()
    return f"TNS-{date_str}-CRT{court_id}-{random_suffix}"

def generate_qr_code(data: str, size: int = 10) -> BytesIO:
    """QR kod yaratish"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=size,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # BytesIO ga saqlash
    img_buffer = BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer

def save_qr_code(data: str, file_path: str, size: int = 10) -> str:
    """QR kodni faylga saqlash"""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    qr_buffer = generate_qr_code(data, size)
    
    with open(file_path, 'wb') as f:
        f.write(qr_buffer.getvalue())
    
    return file_path

def create_ticket_pdf(ticket_data: Dict, file_path: str) -> str:
    """Bilet PDF yaratish"""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Sarlavha
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("🎾 TENNIS KORT BİLETİ", title_style))
    
    # Bilet ma'lumotlari
    data = [
        ['Bilet ID:', ticket_data['ticket_id']],
        ['Mijoz:', f"{ticket_data['user_name']}"],
        ['Telefon:', ticket_data['phone']],
        ['Sana:', ticket_data['date']],
        ['Vaqt:', f"{ticket_data['start_time']} - {ticket_data['end_time']}"],
        ['Kort:', ticket_data['court_name']],
        ['Narx:', f"{ticket_data['amount']:,.0f} so'm"],
        ['To\'lov holati:', ticket_data['payment_status']],
        ['Yaratilgan:', ticket_data['created_at']],
    ]
    
    table = Table(data, colWidths=[2*inch, 4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    
    # QR kod haqida ma'lumot
    qr_info = Paragraph(
        "QR kodni kort kiraverishida ko'rsating:",
        styles['Normal']
    )
    story.append(qr_info)
    story.append(Spacer(1, 10))
    
    # QR kod (bu yerda QR kod rasmini qo'shish kerak)
    # QR kod fayli mavjud bo'lsa, uni qo'shish
    if 'qr_code_path' in ticket_data and os.path.exists(ticket_data['qr_code_path']):
        # QR kodni PDF ga qo'shish logikasi
        pass
    
    # Qoidalar
    rules_text = """
    <b>Muhim eslatmalar:</b><br/>
    • Biletni kort kiraverishida ko'rsating<br/>
    • Kechikish 15 daqiqadan oshsa, bron bekor qilinadi<br/>
    • Raketkalar va to'plar alohida ijaraga beriladi<br/>
    • Ichimlik va ovqat olib kirish taqiqlangan<br/>
    • Qo'llab-quvvatlash: +998 90 123 45 67
    """
    
    rules = Paragraph(rules_text, styles['Normal'])
    story.append(Spacer(1, 30))
    story.append(rules)
    
    doc.build(story)
    return file_path

def calculate_booking_price(
    court_hourly_rate_peak: float,
    court_hourly_rate_offpeak: float,
    duration_hours: float,
    start_time: datetime,
    is_vip: bool = False,
    promo_discount: float = 0.0
) -> Dict[str, float]:
    """Bron narxini hisoblash"""
    
    # Asosiy narxni aniqlash
    hour = start_time.hour
    is_peak = Config.is_peak_time(hour)
    is_weekend = Config.is_weekend(start_time.weekday())
    
    base_rate = court_hourly_rate_peak if is_peak else court_hourly_rate_offpeak
    base_price = base_rate * duration_hours
    
    # Peak vaqt qo'shimchasi
    peak_extra = 0.0
    if is_peak:
        peak_extra = (court_hourly_rate_peak - court_hourly_rate_offpeak) * duration_hours
    
    # Dam olish kuni qo'shimchasi
    weekend_extra = 0.0
    if is_weekend:
        weekend_coefficient = Config.WEEKEND_COEFFICIENT
        weekend_extra = base_price * (weekend_coefficient - 1)
    
    # Jami asosiy narx
    subtotal = base_price + weekend_extra
    
    # Chegirmalar
    discount = 0.0
    
    # VIP chegirma
    if is_vip:
        vip_discount = subtotal * Config.VIP_DISCOUNT
        discount += vip_discount
    
    # Promo kod chegirmasi
    if promo_discount > 0:
        discount += promo_discount
    
    # Xizmat haqi
    service_fee = subtotal * Config.SERVICE_FEE
    
    # Yakuniy narx
    final_amount = subtotal - discount + service_fee
    
    return {
        'base_price': base_price,
        'peak_extra': peak_extra,
        'weekend_extra': weekend_extra,
        'subtotal': subtotal,
        'discount': discount,
        'service_fee': service_fee,
        'final_amount': max(0, final_amount),  # Manfiy bo'lmasligi uchun
        'is_peak': is_peak,
        'is_weekend': is_weekend
    }

def validate_phone_number(phone: str) -> bool:
    """Telefon raqamni tekshirish"""
    # O'zbekiston telefon raqamlari uchun regex
    pattern = r'^(\+998|998|8)?[0-9]{9}$'
    return bool(re.match(pattern, phone.replace(' ', '').replace('-', '')))

def format_phone_number(phone: str) -> str:
    """Telefon raqamni formatlash"""
    # Barcha maxsus belgilarni olib tashlash
    cleaned = re.sub(r'[^\d]', '', phone)
    
    # O'zbekiston kodi qo'shish
    if cleaned.startswith('998'):
        cleaned = '+' + cleaned
    elif cleaned.startswith('8') and len(cleaned) == 9:
        cleaned = '+998' + cleaned[1:]
    elif len(cleaned) == 9:
        cleaned = '+998' + cleaned
    
    return cleaned

def get_uzbekistan_time() -> datetime:
    """O'zbekiston vaqtini olish"""
    tz = pytz.timezone(Config.TIMEZONE)
    return datetime.now(tz)

def is_working_hours(dt: datetime) -> bool:
    """Ish vaqtini tekshirish"""
    hour = dt.hour
    return Config.COURT_OPEN_HOUR <= hour < Config.COURT_CLOSE_HOUR

def get_available_time_slots(date: datetime, court_id: int, 
                           booked_slots: List[Dict]) -> List[Dict]:
    """Mavjud vaqt slotlarini olish"""
    slots = []
    
    # Ish vaqti davomida har soat uchun slot yaratish
    start_hour = Config.COURT_OPEN_HOUR
    end_hour = Config.COURT_CLOSE_HOUR
    
    for hour in range(start_hour, end_hour):
        slot_start = date.replace(hour=hour, minute=0, second=0, microsecond=0)
        slot_end = slot_start + timedelta(hours=1)
        
        # Bu slot band emasligini tekshirish
        is_booked = any(
            slot['start_time'] <= slot_start < slot['end_time'] or
            slot['start_time'] < slot_end <= slot['end_time']
            for slot in booked_slots
        )
        
        if not is_booked:
            slots.append({
                'start_time': slot_start,
                'end_time': slot_end,
                'is_peak': Config.is_peak_time(hour),
                'is_available': True
            })
    
    return slots

def create_booking_summary(booking_data: Dict, lang: str = "uz") -> str:
    """Bron xulosasini yaratish"""
    from localization import get_text
    
    # Kalitlarni xavfsiz olish
    date = booking_data.get('date', '')
    start_time = booking_data.get('start_time_str', '')
    end_time = booking_data.get('end_time_str', '')
    court_name = booking_data.get('court_name', '')
    duration = booking_data.get('duration', 1)
    base_price = booking_data.get('base_price', 0)
    peak_extra = booking_data.get('peak_extra', 0)
    weekend_extra = booking_data.get('weekend_extra', 0)
    discount = booking_data.get('discount', 0)
    service_fee = booking_data.get('service_fee', 0)
    total_amount = booking_data.get('final_amount', booking_data.get('total_amount', 0))
    
    return get_text("booking_details", lang).format(
        date=date,
        start_time=start_time,
        end_time=end_time,
        court_name=court_name,
        duration=duration,
        base_price=base_price,
        peak_extra=peak_extra,
        weekend_extra=weekend_extra,
        discount=discount,
        service_fee=service_fee,
        total_amount=total_amount
    )

def generate_payment_signature(merchant_id: str, amount: int, order_id: str, 
                             secret_key: str) -> str:
    """To'lov imzosini yaratish (Click/Payme uchun)"""
    data = f"{merchant_id}{amount}{order_id}"
    signature = hmac.new(
        secret_key.encode('utf-8'),
        data.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    return signature

def verify_payment_signature(received_signature: str, expected_data: str, 
                           secret_key: str) -> bool:
    """To'lov imzosini tekshirish"""
    expected_signature = hmac.new(
        secret_key.encode('utf-8'),
        expected_data.encode('utf-8'),
        hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(received_signature, expected_signature)

def format_currency(amount: float, currency: str = "UZS") -> str:
    """Valyutani formatlash"""
    if currency == "UZS":
        return f"{amount:,.0f} so'm"
    else:
        return f"{amount:,.2f} {currency}"

def get_booking_status_emoji(status: str) -> str:
    """Bron holati uchun emoji"""
    status_emojis = {
        'pending': '⏳',
        'hold': '🔒',
        'paid': '💳',
        'confirmed': '✅',
        'cancelled': '❌',
        'completed': '✅',
        'no_show': '🚫'
    }
    return status_emojis.get(status, '❓')

def truncate_text(text: str, max_length: int = 50) -> str:
    """Textni qisqartirish"""
    if len(text) <= max_length:
        return text
    return text[:max_length-3] + "..."

def is_valid_date_range(start_date: datetime, end_date: datetime, 
                       max_days: int = 30) -> bool:
    """Sana oralig'ini tekshirish"""
    if start_date >= end_date:
        return False
    
    delta = end_date - start_date
    return delta.days <= max_days

def get_week_dates(date: datetime) -> List[datetime]:
    """Hafta sanalarini olish"""
    start_of_week = date - timedelta(days=date.weekday())
    return [start_of_week + timedelta(days=i) for i in range(7)]

def sanitize_filename(filename: str) -> str:
    """Fayl nomini tozalash"""
    # Xavfli belgilarni olib tashlash
    sanitized = re.sub(r'[^\w\-_\.]', '_', filename)
    return sanitized[:100]  # Uzunlikni cheklash

def create_excel_report(data: List[Dict], headers: List[str], 
                       file_path: str) -> str:
    """Excel hisobotini yaratish"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hisobot"
        
        # Sarlavhalar
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Ma'lumotlar
        for row, item in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=item.get(header, ''))
        
        # Ustunlar kengligini avtomatik sozlash
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
        return file_path
        
    except ImportError:
        raise ImportError("openpyxl kutubxonasi o'rnatilmagan")

def create_ticket_image(ticket_data: Dict, qr_path: str, file_path: str) -> str:
    """Bilet rasmini yaratish"""
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Rasm o'lchamlari - kichikroq va chiroyliroq
    width = 600
    height = 400
    
    # Oq fon yaratish
    img = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(img)
    
    try:
        # Font yuklash - kichikroq o'lchamlar
        title_font = ImageFont.truetype("arial.ttf", 24)
        header_font = ImageFont.truetype("arial.ttf", 18)
        text_font = ImageFont.truetype("arial.ttf", 14)
        small_font = ImageFont.truetype("arial.ttf", 12)
    except:
        # Agar font topilmasa, default font ishlatish
        title_font = ImageFont.load_default()
        header_font = ImageFont.load_default()
        text_font = ImageFont.load_default()
        small_font = ImageFont.load_default()
    
    # Sarlavha
    title = "TENNIS KORT BILETI"
    title_bbox = draw.textbbox((0, 0), title, font=title_font)
    title_width = title_bbox[2] - title_bbox[0]
    title_x = (width - title_width) // 2
    draw.text((title_x, 15), title, fill='black', font=title_font)
    
    # Chiziq chizish
    draw.line([(30, 50), (width - 30, 50)], fill='black', width=2)
    
    # Bilet ma'lumotlari
    y_pos = 70
    
    # Chap tomon - ma'lumotlar
    left_x = 30
    right_x = 320
    
    # Bilet ID
    draw.text((left_x, y_pos), f"Bilet ID:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['ticket_id'], fill='black', font=text_font)
    y_pos += 25
    
    # Kort
    draw.text((left_x, y_pos), f"Kort:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['court_name'], fill='black', font=text_font)
    y_pos += 25
    
    # Sana
    draw.text((left_x, y_pos), f"Sana:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['date'], fill='black', font=text_font)
    y_pos += 25
    
    # Vaqt
    draw.text((left_x, y_pos), f"Vaqt:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), f"{ticket_data['start_time']} - {ticket_data['end_time']}", fill='black', font=text_font)
    y_pos += 25
    
    # Mijoz
    draw.text((left_x, y_pos), f"Mijoz:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['user_name'], fill='black', font=text_font)
    y_pos += 25
    
    # Telefon
    draw.text((left_x, y_pos), f"Telefon:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['phone'], fill='black', font=text_font)
    y_pos += 25
    
    # Summa
    draw.text((left_x, y_pos), f"Summa:", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), f"{ticket_data['amount']:,.0f} so'm", fill='black', font=text_font)
    y_pos += 25
    
    # Bilet olingan
    draw.text((left_x, y_pos), f"Bilet olingan: ", fill='black', font=text_font)
    draw.text((left_x + 80, y_pos), ticket_data['created_at'], fill='black', font=text_font)
    
    # QR kod qo'shish
    if os.path.exists(qr_path):
        qr_img = Image.open(qr_path)
        # QR kod o'lchamini kichraytirish
        qr_size = 150
        qr_img = qr_img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
        
        # QR kodni o'ng tomonga joylash
        qr_x = right_x
        qr_y = 100
        img.paste(qr_img, (qr_x, qr_y))
        
        # QR kod haqida matn
        qr_text = "QR kodni kirish vaqtida ko'rsating"
        qr_text_bbox = draw.textbbox((0, 0), qr_text, font=small_font)
        qr_text_width = qr_text_bbox[2] - qr_text_bbox[0]
        qr_text_x = qr_x + (qr_size - qr_text_width) // 2
        draw.text((qr_text_x, qr_y + qr_size + 10), qr_text, fill='black', font=small_font)
    
    # Pastki qism - qoidalar
    y_pos = qr_y + qr_size + 50
    draw.line([(30, y_pos), (width - 30, y_pos)], fill='black', width=2)
    y_pos += 15
    
    # Qoidalar
    rules = [
        "Ushbu bilet faqat ko'rsatilgan sana va vaqtda amal qiladi.",
        "Biletni boshqa shaxslarga berish taqiqlanadi."
    ]
    
    for rule in rules:
        draw.text((30, y_pos), rule, fill='black', font=small_font)
        y_pos += 20
    
    # Rasmni saqlash
    img.save(file_path, 'PNG', quality=95)
    return file_path

def log_user_action(user_id: int, action: str, details: str = None):
    """Foydalanuvchi amallarini loglash"""
    timestamp = get_uzbekistan_time()
    log_entry = f"[{timestamp}] User {user_id}: {action}"
    if details:
        log_entry += f" - {details}"
    
    # Bu yerda logging kutubxonasidan foydalanish kerak
    print(log_entry)  # Hozircha print, keyinchalik logging.info() ga o'zgartirish
